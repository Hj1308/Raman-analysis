"""
Raman Analyzer — Graphical User Interface
Author: Hoda Jaafari
Run: python app.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os, sys
import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure

# Add src to path
sys.path.insert(0, os.path.dirname(__file__))
from src.loader      import load_spectrum, load_excel_sheets
from src.baseline    import correct_baseline
from src.peak_fitter import fit_all_peaks
from src.analyzer    import analyze, format_report
from src.exporter    import append_csv, save_text_report

# ── Colour palette ────────────────────────────────
BG        = "#1e1e2e"
SURFACE   = "#2a2a3e"
SURFACE2  = "#313145"
ACCENT    = "#4fc3f7"
ACCENT2   = "#81d4fa"
GREEN     = "#69db7c"
ORANGE    = "#ffa94d"
RED       = "#ff6b6b"
TEXT      = "#cdd6f4"
MUTED     = "#7f849c"
BORDER    = "#45475a"

PEAK_COLORS = {
    "D": "#ff6b6b", "G": "#69db7c", "D_prime": "#ffa94d",
    "2D": "#4fc3f7", "DG": "#cc99ff"
}


class RamanApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Raman Spectrum Analyzer  •  Graphene & sp² Carbon Materials")
        self.geometry("1280x820")
        self.minsize(1100, 700)
        self.configure(bg=BG)

        # State
        self.filepath    = tk.StringVar()
        self.laser_nm    = tk.DoubleVar(value=532.0)
        self.baseline_m  = tk.StringVar(value="als")
        self.als_lam     = tk.DoubleVar(value=1e5)
        self.als_p       = tk.DoubleVar(value=0.001)
        self.output_dir  = tk.StringVar(value=os.path.join(os.path.dirname(__file__), "results"))
        self.strict_mode = tk.BooleanVar(value=True)

        # current sample state
        self._wn = None; self._intensity = None
        self._corrected = None; self._baseline = None
        self._peaks = None; self._analysis = None
        self._sample_label = ""

        # multi-sample batch state
        self._batch_samples  = []   # list of dicts from load_excel_sheets
        self._batch_results  = []   # list of (label, peaks, analysis)
        self._batch_idx      = 0

        self._build_ui()
        self._apply_styles()

    # ─────────────────────────────────────────────
    #  UI BUILD
    # ─────────────────────────────────────────────
    def _build_ui(self):
        topbar = tk.Frame(self, bg=SURFACE, height=52)
        topbar.pack(fill="x", side="top")
        topbar.pack_propagate(False)
        tk.Label(topbar, text="🔬  Raman Spectrum Analyzer",
                 bg=SURFACE, fg=ACCENT,
                 font=("Helvetica", 16, "bold")).pack(side="left", padx=20, pady=12)
        tk.Label(topbar, text="Graphene / sp² Carbon Materials",
                 bg=SURFACE, fg=MUTED,
                 font=("Helvetica", 10)).pack(side="left", pady=12)

        main = tk.PanedWindow(self, orient="horizontal",
                              bg=BG, sashwidth=6, sashpad=0, relief="flat")
        main.pack(fill="both", expand=True, padx=8, pady=6)

        left = tk.Frame(main, bg=BG, width=300)
        main.add(left, minsize=260)
        self._build_left(left)

        right = tk.Frame(main, bg=BG)
        main.add(right, minsize=700)
        self._build_right(right)

        self.status_var = tk.StringVar(value="Ready — load a spectrum file to begin")
        statusbar = tk.Frame(self, bg=SURFACE2, height=26)
        statusbar.pack(fill="x", side="bottom")
        statusbar.pack_propagate(False)
        tk.Label(statusbar, textvariable=self.status_var,
                 bg=SURFACE2, fg=MUTED,
                 font=("Helvetica", 9)).pack(side="left", padx=12, pady=4)

    def _build_left(self, parent):
        def section(title):
            f = tk.LabelFrame(parent, text=f"  {title}  ",
                              bg=BG, fg=ACCENT2,
                              font=("Helvetica", 10, "bold"),
                              bd=1, relief="groove",
                              labelanchor="nw")
            f.pack(fill="x", padx=8, pady=(6,2))
            return f

        # ── File ─────────────────────────────────
        fs = section("📂  Input File")
        tk.Button(fs, text="Browse…", command=self._browse_file,
                  bg=ACCENT, fg="#000", font=("Helvetica", 10, "bold"),
                  relief="flat", padx=12, cursor="hand2").pack(fill="x", padx=8, pady=(6,2))
        self._file_label = tk.Label(fs, text="No file selected",
                                    bg=BG, fg=MUTED,
                                    font=("Helvetica", 9),
                                    wraplength=240, justify="left")
        self._file_label.pack(padx=8, pady=(0,2), anchor="w")

        # batch navigator (hidden until an Excel file is loaded)
        self._nav_frame = tk.Frame(fs, bg=BG)
        self._nav_frame.pack(fill="x", padx=8, pady=(0,6))
        tk.Button(self._nav_frame, text="◀", width=3,
                  command=self._prev_sample,
                  bg=SURFACE2, fg=ACCENT, font=("Helvetica", 10, "bold"),
                  relief="flat", cursor="hand2").pack(side="left")
        self._sample_lbl_var = tk.StringVar(value="")
        tk.Label(self._nav_frame, textvariable=self._sample_lbl_var,
                 bg=BG, fg=ACCENT2,
                 font=("Helvetica", 10, "bold"),
                 width=18, anchor="center").pack(side="left", padx=4)
        tk.Button(self._nav_frame, text="▶", width=3,
                  command=self._next_sample,
                  bg=SURFACE2, fg=ACCENT, font=("Helvetica", 10, "bold"),
                  relief="flat", cursor="hand2").pack(side="left")
        self._nav_frame.pack_forget()   # hide initially

        # ── Laser ─────────────────────────────────
        ls = section("🔴  Laser Wavelength")
        laser_frame = tk.Frame(ls, bg=BG)
        laser_frame.pack(fill="x", padx=8, pady=6)
        tk.Label(laser_frame, text="Wavelength (nm):",
                 bg=BG, fg=TEXT, font=("Helvetica", 10)).grid(row=0, column=0, sticky="w")
        tk.Entry(laser_frame, textvariable=self.laser_nm,
                 bg=SURFACE2, fg=ACCENT,
                 font=("Helvetica", 12, "bold"),
                 insertbackground=ACCENT,
                 relief="flat", width=8, justify="center"
                 ).grid(row=0, column=1, padx=(8,0), sticky="w")
        preset_frame = tk.Frame(ls, bg=BG)
        preset_frame.pack(fill="x", padx=8, pady=(0,6))
        tk.Label(preset_frame, text="Presets:", bg=BG, fg=MUTED,
                 font=("Helvetica", 9)).pack(side="left")
        for nm, color in [("488", "#4fc3f7"), ("532", "#69db7c"),
                           ("633", "#ffa94d"), ("785", "#ff6b6b")]:
            tk.Button(preset_frame, text=f"{nm} nm",
                      bg=SURFACE2, fg=color,
                      font=("Helvetica", 9, "bold"),
                      relief="flat", padx=6, cursor="hand2",
                      command=lambda v=nm: self.laser_nm.set(float(v))
                      ).pack(side="left", padx=2)

        # ── Baseline ──────────────────────────────
        bs = section("📉  Baseline Correction")
        bf = tk.Frame(bs, bg=BG); bf.pack(fill="x", padx=8, pady=6)
        tk.Label(bf, text="Method:", bg=BG, fg=TEXT,
                 font=("Helvetica", 10)).grid(row=0, column=0, sticky="w")
        ttk.Combobox(bf, textvariable=self.baseline_m,
                     values=["als", "linear"],
                     state="readonly", width=10
                     ).grid(row=0, column=1, padx=(8,0), sticky="w")
        tk.Label(bf, text="ALS λ (smoothness):", bg=BG, fg=TEXT,
                 font=("Helvetica", 9)).grid(row=1, column=0, sticky="w", pady=(4,0))
        tk.Entry(bf, textvariable=self.als_lam,
                 bg=SURFACE2, fg=TEXT, relief="flat",
                 font=("Helvetica", 9), width=12
                 ).grid(row=1, column=1, padx=(8,0), sticky="w", pady=(4,0))
        tk.Label(bf, text="ALS p (asymmetry):", bg=BG, fg=TEXT,
                 font=("Helvetica", 9)).grid(row=2, column=0, sticky="w", pady=(4,0))
        tk.Entry(bf, textvariable=self.als_p,
                 bg=SURFACE2, fg=TEXT, relief="flat",
                 font=("Helvetica", 9), width=12
                 ).grid(row=2, column=1, padx=(8,0), sticky="w", pady=(4,0))

        # ── Output ────────────────────────────────
        os_frame = section("💾  Output Directory")
        tk.Button(os_frame, text="Choose folder…",
                  command=self._browse_output,
                  bg=SURFACE2, fg=TEXT,
                  font=("Helvetica", 9), relief="flat",
                  padx=8, cursor="hand2").pack(fill="x", padx=8, pady=(6,2))
        self._out_label = tk.Label(os_frame,
                                   text=self.output_dir.get(),
                                   bg=BG, fg=MUTED,
                                   font=("Helvetica", 8),
                                   wraplength=240, justify="left")
        self._out_label.pack(padx=8, pady=(0,6), anchor="w")

        # ── Buttons ───────────────────────────────
        tk.Frame(parent, bg=BG, height=8).pack()
        self._run_btn = tk.Button(parent,
                                  text="▶   RUN ANALYSIS",
                                  command=self._run,
                                  bg=GREEN, fg="#000",
                                  font=("Helvetica", 13, "bold"),
                                  relief="flat", padx=16, pady=12,
                                  cursor="hand2", activebackground="#51cf66")
        self._run_btn.pack(fill="x", padx=8)

        self._export_btn = tk.Button(parent,
                                     text="📊  Export Excel",
                                     command=self._export_excel,
                                     bg=SURFACE2, fg=ACCENT2,
                                     font=("Helvetica", 11, "bold"),
                                     relief="flat", padx=12, pady=8,
                                     cursor="hand2", state="disabled")
        self._export_btn.pack(fill="x", padx=8, pady=(6,0))

        # export ALL button (only visible for Excel batch)
        self._export_all_btn = tk.Button(parent,
                                          text="📦  Export ALL Samples",
                                          command=self._export_all_excel,
                                          bg=SURFACE2, fg=ORANGE,
                                          font=("Helvetica", 10, "bold"),
                                          relief="flat", padx=12, pady=6,
                                          cursor="hand2", state="disabled")
        self._export_all_btn.pack(fill="x", padx=8, pady=(4,0))
        self._export_all_btn.pack_forget()

    def _build_right(self, parent):
        nb = ttk.Notebook(parent)
        nb.pack(fill="both", expand=True)
        self._nb = nb

        tab1 = tk.Frame(nb, bg=BG)
        nb.add(tab1, text="  📈 Spectrum  ")
        self._fig1 = Figure(figsize=(9,5), facecolor="#1a1a2e")
        self._ax1a = self._fig1.add_subplot(211)
        self._ax1b = self._fig1.add_subplot(212)
        canvas1 = FigureCanvasTkAgg(self._fig1, tab1)
        canvas1.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(canvas1, tab1).pack(fill="x")
        self._canvas1 = canvas1

        tab2 = tk.Frame(nb, bg=BG)
        nb.add(tab2, text="  🔍 Peak Fits  ")
        self._fig2 = Figure(figsize=(9,5), facecolor="#1a1a2e")
        canvas2 = FigureCanvasTkAgg(self._fig2, tab2)
        canvas2.get_tk_widget().pack(fill="both", expand=True)
        NavigationToolbar2Tk(canvas2, tab2).pack(fill="x")
        self._canvas2 = canvas2

        tab3 = tk.Frame(nb, bg=BG)
        nb.add(tab3, text="  📋 Results  ")
        self._build_results_tab(tab3)

        tab4 = tk.Frame(nb, bg=BG)
        nb.add(tab4, text="  📄 Report  ")
        self._report_text = scrolledtext.ScrolledText(
            tab4, bg="#0d1117", fg="#c9d1d9",
            font=("Courier", 10), relief="flat",
            insertbackground=ACCENT, state="disabled")
        self._report_text.pack(fill="both", expand=True, padx=4, pady=4)

    def _build_results_tab(self, parent):
        rf = tk.LabelFrame(parent, text="  Intensity Ratios  ",
                           bg=BG, fg=ACCENT2,
                           font=("Helvetica", 10, "bold"), bd=1)
        rf.pack(fill="x", padx=10, pady=8)

        for ci, h in enumerate(["Ratio", "Height-based", "Area-based"]):
            tk.Label(rf, text=h, bg=SURFACE, fg=ACCENT,
                     font=("Helvetica", 10, "bold"),
                     width=20, relief="flat", pady=4
                     ).grid(row=0, column=ci, padx=1, pady=1, sticky="ew")

        self._ratio_vars = {}
        for ri, lbl in enumerate(["ID/IG", "I2D/IG", "ID'/IG", "ID/ID'"]):
            tk.Label(rf, text=lbl, bg=SURFACE2, fg=TEXT,
                     font=("Helvetica", 11, "bold"),
                     width=20, pady=4).grid(row=ri+1, column=0, padx=1, pady=1, sticky="ew")
            h_var = tk.StringVar(value="—")
            a_var = tk.StringVar(value="—")
            self._ratio_vars[lbl] = (h_var, a_var)
            tk.Label(rf, textvariable=h_var, bg=SURFACE2, fg=GREEN,
                     font=("Courier", 11, "bold"),
                     width=20, pady=4).grid(row=ri+1, column=1, padx=1, pady=1, sticky="ew")
            tk.Label(rf, textvariable=a_var, bg=SURFACE2, fg=ACCENT2,
                     font=("Courier", 11, "bold"),
                     width=20, pady=4).grid(row=ri+1, column=2, padx=1, pady=1, sticky="ew")

        pf = tk.LabelFrame(parent, text="  Fitted Peaks  ",
                           bg=BG, fg=ACCENT2,
                           font=("Helvetica", 10, "bold"), bd=1)
        pf.pack(fill="x", padx=10, pady=(0,8))

        for ci, h in enumerate(["Peak", "Center (cm⁻¹)", "FWHM (cm⁻¹)", "Height", "Area", "R²", "Status"]):
            tk.Label(pf, text=h, bg=SURFACE, fg=ACCENT,
                     font=("Helvetica", 9, "bold"),
                     pady=4, relief="flat"
                     ).grid(row=0, column=ci, padx=1, pady=1, sticky="ew")

        self._peak_vars = {}
        for ri, name in enumerate(["D", "G", "D'", "2D", "D+G"]):
            color = list(PEAK_COLORS.values())[ri]
            tk.Label(pf, text=name, bg=SURFACE2, fg=color,
                     font=("Helvetica", 11, "bold"), pady=3
                     ).grid(row=ri+1, column=0, padx=1, pady=1, sticky="ew")
            row_vars = []
            for ci in range(1, 7):
                v = tk.StringVar(value="—")
                row_vars.append(v)
                fg_color = TEXT if ci < 6 else MUTED
                tk.Label(pf, textvariable=v, bg=SURFACE2, fg=fg_color,
                         font=("Courier", 10), pady=3
                         ).grid(row=ri+1, column=ci, padx=1, pady=1, sticky="ew")
            self._peak_vars[name] = row_vars

        sf = tk.LabelFrame(parent, text="  Structural Analysis  ",
                           bg=BG, fg=ACCENT2,
                           font=("Helvetica", 10, "bold"), bd=1)
        sf.pack(fill="x", padx=10, pady=(0,8))
        self._struct_vars = {}
        for ri, lbl in enumerate(["L_D (nm)", "Disorder Stage", "Defect Type", "Estimated Layers"]):
            tk.Label(sf, text=lbl, bg=SURFACE2, fg=TEXT,
                     font=("Helvetica", 10, "bold"),
                     width=22, pady=4, anchor="w"
                     ).grid(row=ri, column=0, padx=(8,1), pady=1, sticky="ew")
            v = tk.StringVar(value="—")
            self._struct_vars[lbl] = v
            tk.Label(sf, textvariable=v, bg=SURFACE2, fg=ORANGE,
                     font=("Helvetica", 10), pady=4, anchor="w"
                     ).grid(row=ri, column=1, padx=(1,8), pady=1, sticky="ew")

    # ─────────────────────────────────────────────
    #  FILE BROWSE
    # ─────────────────────────────────────────────
    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Raman spectrum file",
            filetypes=[
                ("All supported", "*.txt *.csv *.xlsx *.xls"),
                ("Excel workbook", "*.xlsx *.xls"),
                ("Text/CSV files", "*.txt *.csv"),
                ("All files", "*.*"),
            ]
        )
        if not path:
            return
        self.filepath.set(path)
        fname = os.path.basename(path)
        ext   = os.path.splitext(path)[1].lower()

        if ext in (".xlsx", ".xls"):
            # load all sheets immediately so user can navigate
            try:
                samples = load_excel_sheets(path)
                self._batch_samples = samples
                self._batch_results = [None] * len(samples)
                self._batch_idx     = 0
                self._file_label.config(
                    text=f"{fname}  [{len(samples)} sheet(s)]", fg=ACCENT2)
                self._show_batch_nav()
                self._status(f"Excel loaded: {len(samples)} sample sheet(s) — press RUN to analyse")
            except Exception as e:
                messagebox.showerror("Load Error", str(e))
        else:
            self._batch_samples = []
            self._batch_results = []
            self._hide_batch_nav()
            self._file_label.config(text=fname, fg=ACCENT2)
            self._status(f"File loaded: {fname}")

    def _show_batch_nav(self):
        self._nav_frame.pack(fill="x", padx=8, pady=(0,6))
        self._export_all_btn.pack(fill="x", padx=8, pady=(4,0))
        self._update_nav_label()

    def _hide_batch_nav(self):
        self._nav_frame.pack_forget()
        self._export_all_btn.pack_forget()

    def _update_nav_label(self):
        if self._batch_samples:
            s = self._batch_samples[self._batch_idx]
            total = len(self._batch_samples)
            self._sample_lbl_var.set(f"{s['label']}  ({self._batch_idx+1}/{total})")

    def _prev_sample(self):
        if self._batch_samples:
            self._batch_idx = (self._batch_idx - 1) % len(self._batch_samples)
            self._update_nav_label()
            if self._batch_results[self._batch_idx] is not None:
                self._display_cached(self._batch_idx)

    def _next_sample(self):
        if self._batch_samples:
            self._batch_idx = (self._batch_idx + 1) % len(self._batch_samples)
            self._update_nav_label()
            if self._batch_results[self._batch_idx] is not None:
                self._display_cached(self._batch_idx)

    # ─────────────────────────────────────────────
    #  RUN
    # ─────────────────────────────────────────────
    def _browse_output(self):
        d = filedialog.askdirectory(title="Select output directory")
        if d:
            self.output_dir.set(d)
            self._out_label.config(text=d)

    def _status(self, msg, color=None):
        self.status_var.set(msg)

    def _run(self):
        if not self.filepath.get():
            messagebox.showwarning("No file", "Please select a Raman spectrum file first.")
            return
        ext = os.path.splitext(self.filepath.get())[1].lower()
        if ext in (".xlsx", ".xls") and self._batch_samples:
            self._run_batch_current()
        else:
            self._run_single()

    def _run_single(self):
        self._run_btn.config(state="disabled", text="⏳  Analysing…", bg=MUTED)
        self._status("Running analysis…")
        threading.Thread(target=self._run_analysis_single, daemon=True).start()

    def _run_batch_current(self):
        """Analyse the currently selected sheet."""
        self._run_btn.config(state="disabled", text="⏳  Analysing…", bg=MUTED)
        idx = self._batch_idx
        s   = self._batch_samples[idx]
        self._status(f"Analysing '{s['label']}'…")
        threading.Thread(target=self._run_analysis_batch,
                         args=(idx, s), daemon=True).start()

    def _run_analysis_single(self):
        try:
            wn, intensity = load_spectrum(self.filepath.get())
            self._sample_label = os.path.splitext(
                os.path.basename(self.filepath.get()))[0]
            self._finish_analysis(wn, intensity, self._sample_label)
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _run_analysis_batch(self, idx, sample):
        try:
            wn       = sample["wavenumber"]
            intensity= sample["intensity"]
            label    = sample["label"]
            self._finish_analysis(wn, intensity, label, batch_idx=idx)
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _finish_analysis(self, wn, intensity, label, batch_idx=None):
        try:
            self._status(f"[{label}] Baseline correction…")
            lam = float(self.als_lam.get())
            p   = float(self.als_p.get())
            corrected, baseline = correct_baseline(
                wn, intensity, method=self.baseline_m.get(), lam=lam, p=p)

            self._status(f"[{label}] Fitting peaks…")
            laser = float(self.laser_nm.get())
            peaks = fit_all_peaks(wn, corrected, laser_nm=laser)
            analysis = analyze(peaks, laser_nm=laser)

            # cache
            self._wn = wn; self._intensity = intensity
            self._corrected = corrected; self._baseline = baseline
            self._peaks = peaks; self._analysis = analysis
            self._sample_label = label

            if batch_idx is not None:
                self._batch_results[batch_idx] = {
                    "label": label, "wn": wn, "intensity": intensity,
                    "baseline": baseline, "corrected": corrected,
                    "peaks": peaks, "analysis": analysis
                }

            self.after(0, lambda: self._update_ui(
                wn, intensity, baseline, corrected, peaks, analysis, laser, label))
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _display_cached(self, idx):
        r = self._batch_results[idx]
        if r is None:
            return
        laser = float(self.laser_nm.get())
        self._wn = r["wn"]; self._intensity = r["intensity"]
        self._corrected = r["corrected"]; self._baseline = r["baseline"]
        self._peaks = r["peaks"]; self._analysis = r["analysis"]
        self._sample_label = r["label"]
        self._update_ui(r["wn"], r["intensity"], r["baseline"],
                        r["corrected"], r["peaks"], r["analysis"],
                        laser, r["label"])

    # ─────────────────────────────────────────────
    #  UPDATE UI
    # ─────────────────────────────────────────────
    def _update_ui(self, wn, intensity, baseline, corrected, peaks, analysis, laser, label=None):
        fname = label or os.path.basename(self.filepath.get())

        # ── Plot 1: Spectrum ──────────────────────
        self._fig1.clf()
        ax1 = self._fig1.add_subplot(211)
        ax2 = self._fig1.add_subplot(212)
        for ax in [ax1, ax2]:
            ax.set_facecolor("#0d1117")
            ax.tick_params(colors=MUTED, labelsize=8)
            ax.spines['bottom'].set_color(BORDER)
            ax.spines['left'].set_color(BORDER)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

        ax1.plot(wn, intensity,  color="#4fc3f7", lw=1.0, label="Raw spectrum", alpha=0.9)
        ax1.plot(wn, baseline,   color="#ff6b6b", lw=1.5, ls="--", label="ALS baseline", alpha=0.85)
        ax1.fill_between(wn, baseline, intensity, alpha=0.08, color="#4fc3f7")
        ax1.set_ylabel("Intensity (a.u.)", color=TEXT, fontsize=9)
        ax1.set_title(f"{fname}  |  λ = {laser:.0f} nm", color=ACCENT2, fontsize=10, pad=6)
        ax1.legend(facecolor=SURFACE, edgecolor=BORDER, labelcolor=TEXT, fontsize=8)

        ax2.plot(wn, corrected, color="#69db7c", lw=1.2, label="Corrected", alpha=0.9)
        ax2.axhline(0, color=BORDER, lw=0.8, ls="--")
        for key, p in peaks.items():
            if p.found:
                ax2.axvline(p.center, color=PEAK_COLORS.get(key,"gray"),
                            lw=0.8, ls=":", alpha=0.7)
                ax2.text(p.center, corrected.max()*0.92,
                         p.name, color=PEAK_COLORS.get(key,"gray"),
                         fontsize=8, ha="center", fontweight="bold")
        ax2.set_xlabel("Raman Shift (cm⁻¹)", color=TEXT, fontsize=9)
        ax2.set_ylabel("Intensity (a.u.)", color=TEXT, fontsize=9)
        ax2.legend(facecolor=SURFACE, edgecolor=BORDER, labelcolor=TEXT, fontsize=8)
        self._fig1.tight_layout(pad=1.5)
        self._canvas1.draw()

        # ── Plot 2: Individual peaks ──────────────
        self._fig2.clf()
        found_peaks = [(k,p) for k,p in peaks.items() if p.found and len(p.model_x)>0]
        n = len(found_peaks)
        if n > 0:
            for i, (key, p) in enumerate(found_peaks):
                ax = self._fig2.add_subplot(1, n, i+1)
                ax.set_facecolor("#0d1117")
                ax.tick_params(colors=MUTED, labelsize=7)
                for sp in ['top','right']: ax.spines[sp].set_visible(False)
                for sp in ['bottom','left']: ax.spines[sp].set_color(BORDER)
                color = PEAK_COLORS.get(key, "gray")
                mask  = (wn >= p.model_x[0]) & (wn <= p.model_x[-1])
                xd, yd = wn[mask], corrected[mask]
                ax.scatter(xd, yd, s=4, color="#cdd6f4", alpha=0.5, zorder=3)
                ax.plot(p.model_x, p.model_y, color=color, lw=2.0)
                ax.fill_between(p.model_x, p.model_y, alpha=0.3, color=color)
                ax.set_title(
                    f"{p.name}\n{p.center:.1f} cm⁻¹\nFWHM={p.fwhm:.1f}\nR²={p.r_squared:.3f}",
                    color=color, fontsize=8, pad=4)
                ax.set_xlabel("Raman Shift (cm⁻¹)", color=MUTED, fontsize=7)
                if i == 0:
                    ax.set_ylabel("Intensity", color=MUTED, fontsize=7)
        self._fig2.tight_layout(pad=1.2)
        self._canvas2.draw()

        # ── Results tab ───────────────────────────
        import math
        def fmt(v):  return f"{v:.4f}" if not math.isnan(v) else "N/A"
        def fmti(v): return f"{v:.2f}"  if not math.isnan(v) else "N/A"

        self._ratio_vars["ID/IG"][0].set(fmt(analysis.ID_IG_height))
        self._ratio_vars["ID/IG"][1].set(fmt(analysis.ID_IG_area))
        self._ratio_vars["I2D/IG"][0].set(fmt(analysis.I2D_IG_height))
        self._ratio_vars["I2D/IG"][1].set(fmt(analysis.I2D_IG_area))
        self._ratio_vars["ID'/IG"][0].set(fmt(analysis.IDp_IG_height))
        self._ratio_vars["ID'/IG"][1].set("—")
        self._ratio_vars["ID/ID'"][0].set(fmt(analysis.ID_IDp_height))
        self._ratio_vars["ID/ID'"][1].set("—")

        peak_map = {"D":"D","G":"G","D'":"D_prime","2D":"2D","D+G":"DG"}
        for name, row_vars in self._peak_vars.items():
            key = peak_map[name]
            p   = peaks.get(key)
            if p and p.found:
                row_vars[0].set(f"{p.center:.1f}")
                row_vars[1].set(f"{p.fwhm:.1f}")
                row_vars[2].set(f"{p.amplitude:.1f}")
                row_vars[3].set(f"{p.area:.1f}")
                row_vars[4].set(f"{p.r_squared:.3f}")
                row_vars[5].set("✓ Detected")
            else:
                for v in row_vars: v.set("—")
                row_vars[5].set("Not found")

        self._struct_vars["L_D (nm)"].set(fmti(analysis.L_D_nm))
        self._struct_vars["Disorder Stage"].set(analysis.disorder_stage)
        self._struct_vars["Defect Type"].set(analysis.defect_type)
        self._struct_vars["Estimated Layers"].set(analysis.estimated_layers)

        report = format_report(fname, peaks, analysis, laser)
        self._report_text.config(state="normal")
        self._report_text.delete("1.0", "end")
        self._report_text.insert("1.0", report)
        self._report_text.config(state="disabled")

        self._export_btn.config(state="normal", bg=ACCENT, fg="#000")
        if self._batch_samples:
            all_done = all(r is not None for r in self._batch_results)
            self._export_all_btn.config(
                state="normal" if all_done else "disabled")
        self._run_btn.config(state="normal", text="▶   RUN ANALYSIS", bg=GREEN)
        self._nb.select(2)
        self._status(
            f"✓  [{fname}]  ID/IG = {analysis.ID_IG_height:.4f}  |  "
            f"I2D/IG = {analysis.I2D_IG_height:.4f}  |  "
            f"L_D = {analysis.L_D_nm:.2f} nm  |  {analysis.estimated_layers}")

    def _on_error(self, msg):
        self._run_btn.config(state="normal", text="▶   RUN ANALYSIS", bg=GREEN)
        self._status(f"Error: {msg}")
        messagebox.showerror("Analysis Error", msg)

    # ─────────────────────────────────────────────
    #  EXCEL EXPORT  (single current sample)
    # ─────────────────────────────────────────────
    def _export_excel(self):
        if self._analysis is None:
            return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            title="Save Excel report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"raman_{self._sample_label or 'analysis'}.xlsx"
        )
        if not path:
            return
        try:
            self._status("Exporting Excel…")
            self._do_excel_export(path, self._sample_label,
                                   self._wn, self._intensity,
                                   self._baseline, self._corrected,
                                   self._peaks, self._analysis)
            self._status(f"✓  Excel saved: {path}")
            messagebox.showinfo("Exported", f"Excel file saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ─────────────────────────────────────────────
    #  EXCEL EXPORT  (all batch samples)
    # ─────────────────────────────────────────────
    def _export_all_excel(self):
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            title="Save combined Excel report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="raman_all_samples.xlsx"
        )
        if not path:
            return
        try:
            self._status("Exporting all samples…")
            self._do_excel_export_all(path)
            self._status(f"✓  All samples exported: {path}")
            messagebox.showinfo("Exported", f"Combined Excel file saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ─────────────────────────────────────────────
    #  EXCEL WRITER HELPERS
    # ─────────────────────────────────────────────
    def _do_excel_export(self, path, label, wn, intensity, baseline, corrected, peaks, analysis):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import math

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Raman Analysis"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 3

        H  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        HF = PatternFill('solid', fgColor='1F4E79')
        SF = PatternFill('solid', fgColor='2E75B6')
        N  = Font(name='Calibri', size=11)
        B  = Font(name='Calibri', bold=True, size=11, color='1F4E79')
        C  = Alignment(horizontal='center', vertical='center')
        L  = Alignment(horizontal='left',   vertical='center', indent=1)
        laser = float(self.laser_nm.get())

        def brd():
            s = Side(style='thin', color='BDD7EE')
            return Border(left=s, right=s, top=s, bottom=s)

        ws.merge_cells('B2:H2')
        ws['B2'] = 'Raman Spectroscopy Analysis Report'
        ws['B2'].font = Font(name='Calibri', bold=True, color='1F4E79', size=16)
        ws['B2'].alignment = C
        ws.row_dimensions[2].height = 32
        ws.merge_cells('B3:H3')
        ws['B3'] = (f"Sample: {label}  |  Laser: {laser:.0f} nm  |  "
                    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws['B3'].font = Font(name='Calibri', size=10, color='7F7F7F', italic=True)
        ws['B3'].alignment = C
        ws.row_dimensions[4].height = 8

        ws.merge_cells('B5:G5')
        ws['B5'] = 'Intensity Ratios'
        ws['B5'].font = Font(name='Calibri', bold=True, color='1F4E79', size=13)
        ws.row_dimensions[5].height = 22
        for ci, h in enumerate(['Ratio','Height-based','Area-based','Interpretation']):
            c = ws.cell(row=6, column=2+ci, value=h)
            c.font=H; c.fill=HF; c.alignment=C; c.border=brd()
        ws.row_dimensions[6].height = 22

        an = analysis
        def fv(v): return round(v,4) if not math.isnan(v) else 'N/A'
        ratios = [
            ('ID/IG',  fv(an.ID_IG_height),  fv(an.ID_IG_area),
             'Low defect density' if an.ID_IG_height < 0.5
             else 'Moderate defects' if an.ID_IG_height < 1.0
             else 'High defect density'),
            ('I2D/IG', fv(an.I2D_IG_height), fv(an.I2D_IG_area), an.estimated_layers),
            ("ID'/IG", fv(an.IDp_IG_height), 'N/A', 'Intravalley defect indicator'),
            ("ID/ID'", fv(an.ID_IDp_height), 'N/A',
             an.defect_type[:40] if an.defect_type != 'N/A' else 'N/A'),
        ]
        for ri, (param, hv, av, interp) in enumerate(ratios):
            r   = 7+ri
            f2  = PatternFill('solid', fgColor='D6E4F0' if ri%2==0 else 'EBF3FB')
            for c in range(2,7):
                ws.cell(row=r, column=c).fill=f2
                ws.cell(row=r, column=c).border=brd()
            ws.cell(row=r,column=2,value=param).font=B
            ws.cell(row=r,column=2).alignment=L
            ws.cell(row=r,column=3,value=hv).font=N
            ws.cell(row=r,column=3).alignment=C
            ws.cell(row=r,column=4,value=av).font=N
            ws.cell(row=r,column=4).alignment=C
            ws.cell(row=r,column=5,value=interp).font=N
            ws.cell(row=r,column=5).alignment=L
            ws.row_dimensions[r].height=20

        ws.row_dimensions[11].height=8
        ws.merge_cells('B12:H12')
        ws['B12']='Fitted Peak Parameters'
        ws['B12'].font=Font(name='Calibri',bold=True,color='1F4E79',size=13)
        ws.row_dimensions[12].height=22
        for ci, h in enumerate(['Peak','Center (cm⁻¹)','FWHM (cm⁻¹)',
                                  'Height (a.u.)','Area (a.u.)','R²','Status']):
            c=ws.cell(row=13,column=2+ci,value=h)
            c.font=H; c.fill=SF; c.alignment=C; c.border=brd()
        ws.row_dimensions[13].height=22
        for ri,(name,key) in enumerate([('D','D'),('G','G'),
                                          ("D'",'D_prime'),('2D','2D'),('D+G','DG')]):
            r=14+ri; p=peaks.get(key)
            f2=PatternFill('solid',fgColor='D6E4F0' if ri%2==0 else 'EBF3FB')
            for c in range(2,9):
                ws.cell(row=r,column=c).fill=f2
                ws.cell(row=r,column=c).border=brd()
            ws.cell(row=r,column=2,value=name).font=B
            ws.cell(row=r,column=2).alignment=L
            if p and p.found:
                for ci,v in enumerate([round(p.center,2),round(p.fwhm,2),
                                        round(p.amplitude,1),round(p.area,1),
                                        round(p.r_squared,4)]):
                    ws.cell(row=r,column=3+ci,value=v).font=N
                    ws.cell(row=r,column=3+ci).alignment=C
                ws.cell(row=r,column=8,value='Detected ✓').font=Font(
                    name='Calibri',bold=True,size=10,color='375623')
                ws.cell(row=r,column=8).fill=PatternFill('solid',fgColor='E2EFDA')
            else:
                ws.cell(row=r,column=8,value='Not detected').font=Font(
                    name='Calibri',size=10,color='C00000')
                ws.cell(row=r,column=8).fill=PatternFill('solid',fgColor='FCE4D6')
            ws.cell(row=r,column=8).alignment=C
            ws.cell(row=r,column=8).border=brd()
            ws.row_dimensions[r].height=20

        ws.row_dimensions[19].height=8
        ws.merge_cells('B20:H20')
        ws['B20']='Structural Analysis'
        ws['B20'].font=Font(name='Calibri',bold=True,color='1F4E79',size=13)
        ws.row_dimensions[20].height=22
        struct=[('L_D (nm)', f"{an.L_D_nm:.2f}" if not math.isnan(an.L_D_nm) else 'N/A'),
                ('Disorder Stage',an.disorder_stage),
                ('Defect Type',an.defect_type),
                ('Estimated Layers',an.estimated_layers)]
        for ri,(lbl,val) in enumerate(struct):
            r=21+ri
            f2=PatternFill('solid',fgColor='D6E4F0' if ri%2==0 else 'EBF3FB')
            ws.merge_cells(f'D{r}:H{r}')
            for c in [2,3,4]:
                ws.cell(row=r,column=c).fill=f2
                ws.cell(row=r,column=c).border=brd()
            ws.cell(row=r,column=2,value=lbl).font=B
            ws.cell(row=r,column=2).alignment=L
            ws.cell(row=r,column=4,value=val).font=N
            ws.cell(row=r,column=4).alignment=L
            ws.cell(row=r,column=4).border=brd()
            ws.row_dimensions[r].height=20

        # Spectrum Data sheet
        ws2 = wb.create_sheet("Spectrum Data")
        ws2.sheet_view.showGridLines=False
        ws2.column_dimensions['A'].width=3
        ws2.merge_cells('B2:F2')
        ws2['B2']='Baseline-Corrected Spectrum Data'
        ws2['B2'].font=Font(name='Calibri',bold=True,color='1F4E79',size=14)
        ws2['B2'].alignment=C
        ws2.row_dimensions[2].height=28
        for ci, h in enumerate(['Wavenumber (cm⁻¹)','Raw Intensity',
                                  'ALS Baseline','Corrected Intensity','Normalised (0-1)']):
            c=ws2.cell(row=4,column=2+ci,value=h)
            c.font=H; c.fill=HF; c.alignment=C
        ws2.freeze_panes='B5'
        norm = corrected / corrected.max()
        for i,(w,r2,b2,co,n) in enumerate(zip(wn, intensity, baseline, corrected, norm)):
            row=5+i
            for ci,v in enumerate([round(float(w),3),round(float(r2),3),
                                    round(float(b2),3),round(float(co),3),
                                    round(float(n),5)]):
                ws2.cell(row=row,column=2+ci,value=v)
        for c,w2 in [(2,22),(3,20),(4,20),(5,24),(6,16)]:
            ws2.column_dimensions[get_column_letter(c)].width=w2
        for c,w2 in [(2,22),(3,16),(4,16),(5,36),(6,14),(7,14),(8,16)]:
            ws.column_dimensions[get_column_letter(c)].width=w2
        ws.sheet_properties.tabColor='1F4E79'
        ws2.sheet_properties.tabColor='2E75B6'
        wb.save(path)

    def _do_excel_export_all(self, path):
        """Write one sheet per sample into a single workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import math

        wb    = Workbook()
        laser = float(self.laser_nm.get())
        first = True

        H  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        HF = PatternFill('solid', fgColor='1F4E79')
        N  = Font(name='Calibri', size=10)
        B  = Font(name='Calibri', bold=True, size=10, color='1F4E79')
        C  = Alignment(horizontal='center', vertical='center')
        L  = Alignment(horizontal='left',   vertical='center', indent=1)

        def brd():
            s = Side(style='thin', color='BDD7EE')
            return Border(left=s, right=s, top=s, bottom=s)

        for result in self._batch_results:
            if result is None:
                continue
            lbl = result["label"]
            an  = result["analysis"]
            pk  = result["peaks"]

            ws = wb.active if first else wb.create_sheet()
            first = False
            # sheet name: use label (max 31 chars, no forbidden chars)
            safe = lbl[:31].replace("/","-").replace("\\","-").replace("?","").replace("*","").replace("[","").replace("]","")
            ws.title = safe
            ws.sheet_view.showGridLines = False

            # header
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Sample: {lbl}  |  λ={laser:.0f} nm  |  {datetime.now().strftime('%Y-%m-%d')}"
            ws['A1'].font = Font(name='Calibri', bold=True, color='1F4E79', size=12)
            ws['A1'].alignment = C
            ws.row_dimensions[1].height = 24
            ws.row_dimensions[2].height = 6

            # ratios
            def fv(v): return round(v,4) if not math.isnan(v) else 'N/A'
            headers = ['Ratio','Height','Area','Interpretation']
            for ci,h in enumerate(headers):
                c=ws.cell(row=3,column=1+ci,value=h)
                c.font=H; c.fill=HF; c.alignment=C; c.border=brd()
            ratios = [
                ('ID/IG',  fv(an.ID_IG_height),  fv(an.ID_IG_area),
                 'Low' if an.ID_IG_height<0.5 else 'Moderate' if an.ID_IG_height<1 else 'High'),
                ('I2D/IG', fv(an.I2D_IG_height), fv(an.I2D_IG_area), an.estimated_layers),
                ("ID'/IG",fv(an.IDp_IG_height),'N/A','Intravalley'),
                ("ID/ID'",fv(an.ID_IDp_height),'N/A', an.defect_type[:30] if an.defect_type!='N/A' else 'N/A'),
            ]
            for ri,(param,hv,av,interp) in enumerate(ratios):
                r=4+ri
                f2=PatternFill('solid',fgColor='D6E4F0' if ri%2==0 else 'EBF3FB')
                for c2 in range(1,5):
                    ws.cell(row=r,column=c2).fill=f2
                    ws.cell(row=r,column=c2).border=brd()
                ws.cell(row=r,column=1,value=param).font=B; ws.cell(row=r,column=1).alignment=L
                ws.cell(row=r,column=2,value=hv).font=N;   ws.cell(row=r,column=2).alignment=C
                ws.cell(row=r,column=3,value=av).font=N;   ws.cell(row=r,column=3).alignment=C
                ws.cell(row=r,column=4,value=interp).font=N; ws.cell(row=r,column=4).alignment=L
                ws.row_dimensions[r].height=18

            ws.row_dimensions[8].height=6
            for ci,h in enumerate(['Peak','Center','FWHM','Height','Area','R²','Status']):
                c=ws.cell(row=9,column=1+ci,value=h)
                c.font=H; c.fill=PatternFill('solid',fgColor='2E75B6'); c.alignment=C; c.border=brd()
            for ri,(name,key) in enumerate([('D','D'),('G','G'),("D'",'D_prime'),('2D','2D'),('D+G','DG')]):
                r=10+ri; p=pk.get(key)
                f2=PatternFill('solid',fgColor='D6E4F0' if ri%2==0 else 'EBF3FB')
                for c2 in range(1,8):
                    ws.cell(row=r,column=c2).fill=f2; ws.cell(row=r,column=c2).border=brd()
                ws.cell(row=r,column=1,value=name).font=B; ws.cell(row=r,column=1).alignment=L
                if p and p.found:
                    for ci2,v in enumerate([round(p.center,2),round(p.fwhm,2),
                                            round(p.amplitude,1),round(p.area,1),round(p.r_squared,4)]):
                        ws.cell(row=r,column=2+ci2,value=v).font=N
                        ws.cell(row=r,column=2+ci2).alignment=C
                    ws.cell(row=r,column=7,value='✓').font=Font(name='Calibri',bold=True,size=10,color='375623')
                    ws.cell(row=r,column=7).fill=PatternFill('solid',fgColor='E2EFDA')
                else:
                    ws.cell(row=r,column=7,value='—').font=N
                ws.cell(row=r,column=7).alignment=C; ws.cell(row=r,column=7).border=brd()
                ws.row_dimensions[r].height=18

            for c2,w2 in [(1,12),(2,14),(3,12),(4,14),(5,14),(6,10),(7,14)]:
                ws.column_dimensions[get_column_letter(c2)].width=w2

        wb.save(path)

    # ─────────────────────────────────────────────
    #  STYLES
    # ─────────────────────────────────────────────
    def _apply_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook",        background=BG,      borderwidth=0)
        style.configure("TNotebook.Tab",    background=SURFACE, foreground=MUTED,
                        padding=[12,6],     font=("Helvetica",10,"bold"))
        style.map("TNotebook.Tab",
                  background=[("selected", SURFACE2)],
                  foreground=[("selected", ACCENT)])
        style.configure("TCombobox",
                        fieldbackground=SURFACE2, background=SURFACE2,
                        foreground=TEXT, selectbackground=ACCENT)


if __name__ == "__main__":
    app = RamanApp()
    app.mainloop()
