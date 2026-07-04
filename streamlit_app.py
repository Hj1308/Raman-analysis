"""
Raman Spectrum Analyzer — Multi-Material Streamlit App
Author: Hoda Jaafari
Run:    streamlit run streamlit_app.py

Supported materials:
  - Graphene / sp2 carbon (D, G, D', 2D, D+G, D*)
  - MoS2, WS2, MoSe2, WSe2, MoTe2  (TMDs)
  - h-BN
  - Black Phosphorus
  - MXene

v2.4 / v2.5 UI sync:
  - D* band (I_D*/I_G + C/O note)             [Lee 2021]
  - B-doping fingerprint flag                  [Kim 2012]
  - Fitting uncertainty (center ± σ, FWHM ± σ)[Feature #3]
  - Doping level estimator (n/p + carrier density) [Pisana 2007]
  - Refined stage boundary (FWHM(G) + A_D/A_G)[Wu 2018]
  - Batch statistics + ratio heatmap           [Roadmap #10]
"""

import io
import os
import re
import sys
import math
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from src.loader      import load_spectrum
from src.baseline    import correct_baseline

# Cosmic-ray removal from the preprocessing adapter layer (RamanSPy's
# Whitaker–Hayes when installed, modified z-score fallback otherwise).
# Optional: if the module is missing the app runs exactly as before.
try:
    from preprocessing import despike as _despike, Spectrum as _PreSpectrum
    _HAS_DESPIKE = True
except Exception:
    _HAS_DESPIKE = False
from src.peak_fitter import fit_all_peaks
from src.analyzer    import analyze, format_report

# ══════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════
SHEET_FORBIDDEN = re.compile(r'[\\/:?*\[\]]')
SHEET_MAX_LEN   = 28

MATERIAL_GROUPS = {
    "Graphene / sp² Carbon": [
        "Graphene", "Reduced Graphene Oxide (rGO)",
        "Graphene Oxide (GO)", "Carbon Nanotubes (CNT)",
        "Amorphous Carbon", "Graphite", "N-doped Graphene",
        "Other sp² Carbon"
    ],
    "TMD — Molybdenum": ["MoS2", "MoSe2", "MoTe2", "MoO2"],
    "TMD — Tungsten":   ["WS2", "WSe2", "WTe2"],
    "TMD — Other":      ["NbSe2", "TaS2", "TiSe2", "ReS2", "ReSe2"],
    "Hexagonal Boron Nitride": ["h-BN", "BN nanosheet"],
    "Black Phosphorus / Phosphorene": ["Black Phosphorus", "Phosphorene"],
    "MXene": ["Ti3C2Tx", "Ti2CTx", "V2CTx", "Nb2CTx"],
}

MATERIAL_PEAK_WINDOWS = {
    "graphene": {
        "D*": (1080, 1230), "D":  (1270, 1450), "G": (1500, 1600),
        "D'": (1610, 1680), "2D": (2580, 2780), "D+G": (2850, 2960),
    },
    "mos2":  {"E2g": (370, 395),  "A1g": (398, 420)},
    "ws2":   {"E2g": (345, 365),  "A1g": (410, 430), "2LA": (340, 365)},
    "mose2": {"E2g": (280, 295),  "A1g": (235, 250)},
    "wse2":  {"A1g_E2g": (243, 258), "B2g": (302, 316)},
    "mote2": {"A1g": (168, 178),  "E2g": (230, 242)},
    "hbn":   {"E2g": (1355, 1385)},
    "bp":    {"Ag1": (355, 370),  "B2g": (430, 445), "Ag2": (458, 475)},
    "mxene": {"D":  (1270, 1450), "G": (1500, 1600)},
}

# v2.4: D* colour added, D' key fixed
PEAK_COLORS_GRAPHENE = {
    "D_star": "#e8b4f8",
    "D":      "#ff6b6b",
    "G":      "#69db7c",
    "D_prime":"#ffa94d",
    "D'":     "#ffa94d",
    "2D":     "#4fc3f7",
    "DG":     "#cc99ff",
    "D+G":    "#cc99ff",
}
PEAK_COLORS_TMD = {
    "E2g": "#4fc3f7", "A1g": "#ff6b6b", "2LA": "#ffa94d",
    "B2g": "#cc99ff", "Ag1": "#69db7c", "Ag2": "#4fc3f7",
    "A1g_E2g": "#ff6b6b",
}

# Ratios shown in batch heatmap (must be numeric-castable)
HEATMAP_RATIO_KEYS = [
    "ID/IG (height)", "ID/IG (area)",
    "I2D/IG (height)", "I2D/IG (area)",
    "ID'/IG (height)", "ID/ID' (height)",
    "ID*/IG (height)", "L_D (nm)",
]


# ══════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════
def validate_sample_name(name: str) -> tuple[bool, str]:
    name = name.strip()
    if not name:
        return False, "Sample name cannot be empty."
    if SHEET_FORBIDDEN.search(name):
        return False, "Name contains forbidden characters: \\ / : ? * [ ]  — please remove them."
    if len(name) > SHEET_MAX_LEN:
        return False, f"Name is too long ({len(name)} chars). Max {SHEET_MAX_LEN} characters."
    return True, ""


def _material_key(group: str, material: str) -> str:
    s = (group + " " + material).lower()
    if "mxene" in s:     return "mxene"
    if any(x in s for x in ["graphene", "sp", "carbon", "rgo",
                              "go", "cnt", "graphite", "nanotube"]):
        return "graphene"
    if "mos" in s:        return "mos2"
    if "ws2" in s or ("ws" in s and "2" in s): return "ws2"
    if "mose" in s:       return "mose2"
    if "wse" in s:        return "wse2"
    if "mote" in s:       return "mote2"
    if "bn" in s:         return "hbn"
    if "phosphor" in s:   return "bp"
    return "graphene"


def _fv(v, fmt=".4f"):
    try:
        return format(float(v), fmt) if not math.isnan(float(v)) else "N/A"
    except Exception:
        return "N/A"


def _make_peak_colors(mat_key):
    return PEAK_COLORS_GRAPHENE if mat_key in ("graphene", "mxene") else PEAK_COLORS_TMD


def _parse_sheet_data(xl, sheet_name: str) -> tuple:
    """
    Robustly extract (wavenumber, intensity) arrays from a template sheet.

    Template layout:
        Row 1 (index 0): label header  "Sample Name ..."
        Row 2 (index 1): sample name value  (yellow cell B2)
        Row 3 (index 2): column headers  "Wavenumber (cm-1)" / "Intensity (a.u.)"
        Row 4+ (index 3+): numeric data in columns B and C (index 1 and 2)

    Fallback: scan every row pair (col B, col C) for the first fully-numeric
    pair, then treat everything from that row onward as data.
    """
    df = xl.parse(sheet_name, header=None)

    # --- Strategy 1: template format (data from row index 3, cols 1 & 2) ---
    if df.shape[0] > 3 and df.shape[1] > 2:
        candidate = df.iloc[3:, [1, 2]].copy()
        candidate.columns = ["wavenumber", "intensity"]
        candidate = candidate.apply(pd.to_numeric, errors="coerce").dropna()
        if len(candidate) >= 10:
            return candidate["wavenumber"].values, candidate["intensity"].values

    # --- Strategy 2: scan all rows for first numeric pair ---
    for start in range(df.shape[0]):
        for col_pair in [(1, 2), (0, 1)]:
            if df.shape[1] > col_pair[1]:
                chunk = df.iloc[start:, list(col_pair)].copy()
                chunk.columns = ["wavenumber", "intensity"]
                chunk = chunk.apply(pd.to_numeric, errors="coerce").dropna()
                if len(chunk) >= 10:
                    return chunk["wavenumber"].values, chunk["intensity"].values

    raise ValueError(
        f"Sheet '{sheet_name}': could not find numeric wavenumber/intensity data. "
        "Make sure data starts from row 4 in columns B and C."
    )


# ══════════════════════════════════════════════════════════
#  PEAK FITTING — GENERIC (TMD / h-BN / BP)
# ══════════════════════════════════════════════════════════
from scipy.optimize import curve_fit


def _lorentzian(x, center, amplitude, gamma):
    return amplitude / (np.pi * gamma * (1.0 + ((x - center) / gamma) ** 2))


def fit_single_peak(wn, intensity, lo, hi, name):
    mask = (wn >= lo) & (wn <= hi)
    xd, yd = wn[mask], intensity[mask]
    result = {"name": name, "found": False, "center": np.nan,
              "amplitude": np.nan, "fwhm": np.nan, "area": np.nan,
              "r2": np.nan, "x": xd, "y_fit": np.zeros_like(xd),
              "center_stderr": None, "fwhm_stderr": None}
    if len(xd) < 5 or yd.max() < 1:
        return result
    try:
        c0  = xd[np.argmax(yd)]
        g0  = (hi - lo) / 8.0
        a0  = yd.max() * np.pi * g0
        popt, pcov = curve_fit(
            _lorentzian, xd, yd,
            p0=[c0, a0, g0],
            bounds=([lo, 0, 0.5], [hi, np.inf, (hi - lo) / 2])
        )
        y_fit = _lorentzian(xd, *popt)
        ss_res = np.sum((yd - y_fit) ** 2)
        ss_tot = np.sum((yd - yd.mean()) ** 2)
        r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
        # stderr from pcov
        c_std = float(np.sqrt(pcov[0, 0])) if np.isfinite(pcov[0, 0]) and pcov[0, 0] >= 0 else None
        g_std = float(np.sqrt(pcov[2, 2])) if np.isfinite(pcov[2, 2]) and pcov[2, 2] >= 0 else None
        result.update({
            "found": r2 > 0.65,
            "center": popt[0],
            "amplitude": popt[1] / (np.pi * popt[2]),
            "fwhm": 2.0 * popt[2],
            "area": popt[1],
            "r2": r2,
            "y_fit": y_fit,
            "center_stderr": c_std,
            "fwhm_stderr": (2.0 * g_std) if g_std is not None else None,
        })
    except Exception:
        pass
    return result


def fit_material_peaks(wn, intensity, mat_key):
    windows = MATERIAL_PEAK_WINDOWS.get(mat_key, MATERIAL_PEAK_WINDOWS["graphene"])
    return {name: fit_single_peak(wn, intensity, lo, hi, name)
            for name, (lo, hi) in windows.items()}


# ══════════════════════════════════════════════════════════
#  MATERIAL ANALYSIS
# ══════════════════════════════════════════════════════════
def analyze_graphene(peaks, laser_nm):
    """
    Build the analysis dict surfaced in the UI.
    v2.4/v2.5: includes D*, B-doping, doping estimator, refined stage.
    """
    result = analyze(peaks, laser_nm=laser_nm)
    out = {
        # ── Core ratios ────────────────────────────────
        "ID/IG (height)":      _fv(result.ID_IG_height),
        "ID/IG (area)":        _fv(result.ID_IG_area),
        "I2D/IG (height)":     _fv(result.I2D_IG_height),
        "I2D/IG (area)":       _fv(result.I2D_IG_area),
        "ID'/IG (height)":     _fv(result.IDp_IG_height),
        "ID/ID' (height)":     _fv(result.ID_IDp_height),
        # ── v2.4: D* band ─────────────────────────────
        "ID*/IG (height)":     _fv(result.IDstar_IG_height),
        "D* C/O Note":         result.dstar_co_note if result.dstar_co_note else "—",
        # ── v2.4: B-doping fingerprint ─────────────────
        "⚠ B-Doping Flag":    ("YES ← " + result.boron_doping_note)
                                if result.boron_doping_flag else "No",
        # ── v2.5: doping estimator ─────────────────────
        "Doping Type":         result.doping_type,
        "|n| (cm⁻²)":         _fv(result.carrier_density_cm2, ".3e"),
        "Doping Note":         result.doping_note if result.doping_note else "—",
        # ── v2.5: refined stage boundary ──────────────
        "Stage Refined":       result.stage_refined,
        "Stage Refined Note":  result.stage_refined_note if result.stage_refined_note else "—",
        # ── Structural ────────────────────────────────
        "L_D (nm)":            _fv(result.L_D_nm, ".2f"),
        "L_D Note":            result.L_D_note if result.L_D_note else "—",
        "Disorder Stage":      result.disorder_stage,
        "Defect Type":         result.defect_type,
        "Estimated Layers":    result.estimated_layers,
    }
    return out


def analyze_tmd(peaks, mat_key, laser_nm):
    res = {}
    if mat_key == "mos2":
        e2g, a1g = peaks.get("E2g"), peaks.get("A1g")
        if e2g and a1g and e2g["found"] and a1g["found"]:
            dw = a1g["center"] - e2g["center"]
            res["E2g center (cm-1)"]  = f"{e2g['center']:.1f}"
            res["A1g center (cm-1)"]  = f"{a1g['center']:.1f}"
            res["Delta-omega (A1g-E2g)"] = f"{dw:.1f} cm-1"
            if   dw < 19:   layers = "Monolayer"
            elif dw < 21.5: layers = "Bilayer"
            elif dw < 23:   layers = "Trilayer"
            else:           layers = "Bulk / thick film"
            res["Estimated Layers"] = layers
            res["A1g FWHM (cm-1)"]  = f"{a1g['fwhm']:.1f}"
            if e2g["amplitude"] > 0:
                res["A1g/E2g height ratio"] = f"{a1g['amplitude']/e2g['amplitude']:.3f}"
    elif mat_key == "ws2":
        e2g, a1g, la2 = peaks.get("E2g"), peaks.get("A1g"), peaks.get("2LA")
        if e2g and a1g and e2g["found"] and a1g["found"]:
            dw = a1g["center"] - e2g["center"]
            res["E2g center (cm-1)"]  = f"{e2g['center']:.1f}"
            res["A1g center (cm-1)"]  = f"{a1g['center']:.1f}"
            res["Delta-omega"]        = f"{dw:.1f} cm-1"
            res["Estimated Layers"]   = "Monolayer" if dw < 63 else ("Bilayer" if dw < 66 else "Few-layer / bulk")
        if la2 and la2["found"] and a1g and a1g["found"] and a1g["amplitude"] > 0:
            res["2LA/A1g (defect indicator)"] = f"{la2['amplitude']/a1g['amplitude']:.3f}"
    elif mat_key in ("mose2", "wse2"):
        e2g = peaks.get("E2g") or peaks.get("A1g_E2g")
        if e2g and e2g["found"]:
            res["Main peak center (cm-1)"] = f"{e2g['center']:.1f}"
            res["Main peak FWHM (cm-1)"]   = f"{e2g['fwhm']:.1f}"
        b2g = peaks.get("B2g")
        if b2g and b2g["found"]:
            res["B2g center (cm-1)"] = f"{b2g['center']:.1f}"
    elif mat_key == "mote2":
        a1g, e2g = peaks.get("A1g"), peaks.get("E2g")
        if a1g and a1g["found"]: res["A1g center (cm-1)"] = f"{a1g['center']:.1f}"
        if e2g and e2g["found"]: res["E2g center (cm-1)"] = f"{e2g['center']:.1f}"
        if a1g and e2g and a1g["found"] and e2g["found"]:
            res["Phase"] = "2H (semiconducting)" if a1g["center"] > 170 else "1T' (metallic)"
    return res


def analyze_hbn(peaks):
    e2g = peaks.get("E2g")
    if e2g and e2g["found"]:
        return {
            "E2g center (cm-1)": f"{e2g['center']:.1f}",
            "E2g FWHM (cm-1)":   f"{e2g['fwhm']:.1f}",
            "Crystallinity note": "FWHM<10 cm-1: high; >20 cm-1: defective",
        }
    return {}


def analyze_bp(peaks):
    res = {}
    for pk in ["Ag1", "B2g", "Ag2"]:
        p = peaks.get(pk)
        if p and p["found"]:
            res[f"{pk} center (cm-1)"] = f"{p['center']:.1f}"
            res[f"{pk} FWHM (cm-1)"]  = f"{p['fwhm']:.1f}"
    ag1, ag2 = peaks.get("Ag1"), peaks.get("Ag2")
    if ag1 and ag2 and ag1["found"] and ag2["found"] and ag1["amplitude"] > 0:
        res["Ag2/Ag1 ratio"] = f"{ag2['amplitude']/ag1['amplitude']:.3f}"
    return res


def run_analysis(wn, intensity, mat_key, group, laser_nm,
                 baseline_method, als_lam, als_p, material_label=None):
    # material_label (e.g. "Reduced Graphene Oxide (rGO)") enables:
    #   * 'auto' baseline -> asPLS for fluorescent materials (GO/rGO/g-C3N4)
    #   * adaptive lineshape -> pseudo-Voigt global D/G/D' fit when disordered
    # Step 0: cosmic-ray removal (spikes are common in real GO/rGO spectra and
    # corrupt both the baseline fit and peak amplitudes if left in).
    if _HAS_DESPIKE:
        try:
            _clean = _despike(_PreSpectrum(wn, intensity))
            wn, intensity = _clean.wavenumber, _clean.intensity
        except Exception:
            pass  # never let preprocessing kill the analysis
    corrected, baseline_arr = correct_baseline(
        wn, intensity, method=baseline_method, lam=als_lam, p=als_p,
        material=material_label,
    )
    validation_report = None
    if mat_key in ("graphene", "mxene"):
        peaks_raw  = fit_all_peaks(wn, corrected, laser_nm=laser_nm,
                                   adaptive_lineshape="auto",
                                   material=material_label)
        peaks_dict = {}
        for k, p in peaks_raw.items():
            peaks_dict[k] = {
                "name":           p.name,
                "found":          p.found,
                "center":         p.center,
                "amplitude":      p.amplitude,
                "fwhm":           p.fwhm,
                "area":           p.area,
                "r2":             p.r_squared,
                "x":              p.model_x,
                "y_fit":          p.model_y,
                "is_split_2D":    getattr(p, "is_split_2D",    False),
                "is_deconvolved": getattr(p, "is_deconvolved", False),
                # v2.4 Feature #3: fitting uncertainty
                "center_stderr":  getattr(p, "center_stderr", None),
                "fwhm_stderr":    getattr(p, "fwhm_stderr",   None),
            }
        analysis = analyze_graphene(peaks_raw, laser_nm)
        # Post-fit quality control (pinned bands, missing cores, literature
        # cross-check). Non-fatal: any import/runtime issue is swallowed so
        # the app never breaks because of QC itself.
        try:
            from src.validation import validate
            validation_report = validate(peaks_raw, None, laser_nm=laser_nm)
        except Exception:
            validation_report = None
    else:
        peaks_dict = fit_material_peaks(wn, corrected, mat_key)
        if mat_key in ("mos2", "ws2", "mose2", "wse2", "mote2"):
            analysis = analyze_tmd(peaks_dict, mat_key, laser_nm)
        elif mat_key == "hbn": analysis = analyze_hbn(peaks_dict)
        elif mat_key == "bp":  analysis = analyze_bp(peaks_dict)
        else: analysis = {}
    return peaks_dict, analysis, corrected, baseline_arr, validation_report


# ══════════════════════════════════════════════════════════
#  BATCH STATISTICS HELPERS  (Roadmap #10)
# ══════════════════════════════════════════════════════════
def _build_numeric_df(samples_results, ratio_keys=None):
    """
    Build a DataFrame of numeric ratio values across all samples.
    Rows = samples, columns = ratio keys that could be cast to float.
    """
    if ratio_keys is None:
        ratio_keys = HEATMAP_RATIO_KEYS
    rows = []
    for sr in samples_results:
        row = {"Sample": sr["name"]}
        for k in ratio_keys:
            v = sr["analysis"].get(k, "N/A")
            try:
                row[k] = float(v)
            except (ValueError, TypeError):
                row[k] = np.nan
        rows.append(row)
    df = pd.DataFrame(rows).set_index("Sample")
    # Drop columns that are ALL NaN
    df = df.dropna(axis=1, how="all")
    return df


def _ratio_heatmap(df_num: pd.DataFrame) -> go.Figure:
    """Plotly annotated heatmap of ratio values."""
    z      = df_num.values.astype(float)
    labels = df_num.columns.tolist()
    rows   = df_num.index.tolist()

    # Normalise each column (0→1) for colour scale; keep raw for text
    z_norm = np.zeros_like(z)
    for ci in range(z.shape[1]):
        col = z[:, ci]
        valid = col[~np.isnan(col)]
        if len(valid) > 1 and (valid.max() - valid.min()) > 0:
            z_norm[:, ci] = (col - valid.min()) / (valid.max() - valid.min())
        else:
            z_norm[:, ci] = 0.5

    # Annotation text: raw value (2 decimal places) or "N/A"
    text = []
    for ri in range(z.shape[0]):
        row_txt = []
        for ci in range(z.shape[1]):
            row_txt.append(f"{z[ri, ci]:.3f}" if not np.isnan(z[ri, ci]) else "N/A")
        text.append(row_txt)

    fig = go.Figure(go.Heatmap(
        z=z_norm,
        x=labels,
        y=rows,
        text=text,
        texttemplate="%{text}",
        colorscale="RdYlGn",
        showscale=True,
        colorbar=dict(title="Normalised", tickvals=[0, 0.5, 1],
                      ticktext=["Low", "Mid", "High"]),
        hoverongaps=False,
    ))
    fig.update_layout(
        template="plotly_dark",
        height=max(280, 80 + 50 * len(rows)),
        xaxis=dict(side="top", tickangle=-30),
        yaxis=dict(autorange="reversed"),
        margin=dict(t=120, b=40, l=120, r=40),
        title=dict(text="Raman Ratio Heatmap — batch overview", font=dict(size=13)),
    )
    return fig


def _ratio_bar_chart(df_num: pd.DataFrame, ratio_key: str) -> go.Figure:
    """Bar chart for a single ratio across all samples."""
    palette = ["#4fc3f7","#69db7c","#ff6b6b","#ffa94d",
               "#cc99ff","#ffd43b","#a9e34b","#f783ac"]
    fig = go.Figure()
    vals = df_num[ratio_key].values if ratio_key in df_num.columns else np.array([])
    names = df_num.index.tolist()
    for i, (n, v) in enumerate(zip(names, vals)):
        fig.add_trace(go.Bar(
            x=[n], y=[v] if not np.isnan(v) else [0],
            name=n,
            marker_color=palette[i % len(palette)],
            showlegend=False,
        ))
    fig.update_layout(
        template="plotly_dark", height=320,
        title=dict(text=ratio_key, font=dict(size=12)),
        xaxis_title="Sample",
        yaxis_title=ratio_key,
        margin=dict(t=50, b=60, l=60, r=20),
    )
    return fig


def _numeric_df_to_csv(df_num: pd.DataFrame) -> bytes:
    return df_num.reset_index().to_csv(index=False).encode("utf-8")


# ══════════════════════════════════════════════════════════
#  EXCEL TEMPLATE
# ══════════════════════════════════════════════════════════
def make_template(n_samples: int) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    H  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HF = PatternFill("solid", fgColor="1F4E79")
    YF = PatternFill("solid", fgColor="FFF2CC")
    BF = Font(name="Calibri", bold=True, size=11, color="7F6000")
    C  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_inst = wb.create_sheet("READ ME FIRST", 0)
    ws_inst.sheet_properties.tabColor = "FF0000"
    ws_inst.sheet_view.showGridLines   = False
    ws_inst.column_dimensions["A"].width = 3
    ws_inst.column_dimensions["B"].width = 60
    ws_inst.row_dimensions[2].height   = 30

    ws_inst.merge_cells("B2:C2")
    ws_inst["B2"] = "HOW TO FILL THIS TEMPLATE"
    ws_inst["B2"].font      = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws_inst["B2"].alignment = C
    ws_inst["B2"].fill      = PatternFill("solid", fgColor="D6E4F0")

    instructions = [
        ("Step 1",
         "Go to each sample sheet (Sample_1, Sample_2, …)."),
        ("Step 2",
         "In cell B2 (yellow), type your sample name.\n"
         "RULES for sample name:\n"
         "  • No forbidden characters:  / \\ : ? * [ ]\n"
         "  • Maximum 28 characters\n"
         "  • The name you enter here will be used as the label in all plots and the Excel report."),
        ("Step 3",
         "Starting from row 4, fill column B with Wavenumber (cm-1) "
         "and column C with Intensity (a.u.).\n"
         "Do NOT change row 3 (headers)."),
        ("Step 4",
         "Save the file and upload it to the Raman Analyzer app."),
        ("Note",
         "You can rename the sheet tabs — but the same naming rules apply "
         "(no / \\ : ? * [ ])."),
    ]
    for ri, (step, text) in enumerate(instructions):
        r = 4 + ri * 2
        ws_inst.cell(row=r, column=2, value=step).font = Font(
            name="Calibri", bold=True, size=11, color="C00000")
        ws_inst.cell(row=r + 1, column=2, value=text).font = Font(
            name="Calibri", size=11)
        ws_inst.cell(row=r + 1, column=2).alignment = Alignment(
            wrap_text=True, vertical="top")
        ws_inst.row_dimensions[r + 1].height = 70

    for i in range(1, n_samples + 1):
        ws = wb.create_sheet(title=f"Sample_{i}")
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "2E75B6"

        ws.merge_cells("B1:C1")
        ws["B1"] = "Sample Name  (no / \\ : ? * [ ] — max 28 chars)"
        ws["B1"].font      = Font(name="Calibri", bold=True, size=10, color="7F6000")
        ws["B1"].fill      = PatternFill("solid", fgColor="FFF2CC")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        ws["B2"] = f"Sample_{i}"
        ws["B2"].font      = BF
        ws["B2"].fill      = YF
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 22
        for col, header in enumerate(["Wavenumber (cm-1)", "Intensity (a.u.)"], start=2):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = H; cell.fill = HF; cell.alignment = C

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════
#  EXCEL REPORT EXPORT
# ══════════════════════════════════════════════════════════
def build_excel_report(samples_results, laser_nm):
    wb  = Workbook()
    wb.remove(wb.active)
    H  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HF = PatternFill("solid", fgColor="1F4E79")
    SF = PatternFill("solid", fgColor="2E75B6")
    N  = Font(name="Calibri", size=11)
    B  = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    C  = Alignment(horizontal="center", vertical="center")
    L  = Alignment(horizontal="left",   vertical="center", indent=1)

    def brd():
        s = Side(style="thin", color="BDD7EE")
        return Border(left=s, right=s, top=s, bottom=s)

    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.sheet_properties.tabColor = "1F4E79"
    ws_sum.column_dimensions["A"].width = 3
    ws_sum.merge_cells("B2:J2")
    ws_sum["B2"] = "Raman Spectroscopy — Multi-Sample Analysis Report"
    ws_sum["B2"].font      = Font(name="Calibri", bold=True, color="1F4E79", size=16)
    ws_sum["B2"].alignment = C
    ws_sum.row_dimensions[2].height = 32
    ws_sum.merge_cells("B3:J3")
    ws_sum["B3"] = (
        f"Laser: {laser_nm:.0f} nm  |  Samples: {len(samples_results)}  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  v2.5"
    )
    ws_sum["B3"].font = Font(name="Calibri", size=10, color="7F7F7F", italic=True)
    ws_sum["B3"].alignment = C

    all_params, seen = [], set()
    for sr in samples_results:
        for k in sr["analysis"]:
            if k not in seen:
                all_params.append(k); seen.add(k)

    row = 5
    headers = ["Sample Name", "Material", "Group"] + all_params
    for ci, h in enumerate(headers):
        cell = ws_sum.cell(row=row, column=2 + ci, value=h)
        cell.font = H; cell.fill = HF; cell.alignment = C; cell.border = brd()
    ws_sum.row_dimensions[row].height = 22

    for ri, sr in enumerate(samples_results):
        r   = row + 1 + ri
        alt = PatternFill("solid", fgColor="D6E4F0" if ri % 2 == 0 else "EBF3FB")
        for ci in range(len(headers)):
            ws_sum.cell(row=r, column=2 + ci).fill   = alt
            ws_sum.cell(row=r, column=2 + ci).border = brd()
        ws_sum.cell(row=r, column=2, value=sr["name"]).font = B
        ws_sum.cell(row=r, column=2).alignment = L
        ws_sum.cell(row=r, column=3, value=sr["material"]).font = N
        ws_sum.cell(row=r, column=3).alignment = C
        ws_sum.cell(row=r, column=4, value=sr["group"]).font = N
        ws_sum.cell(row=r, column=4).alignment = C
        for ci, param in enumerate(all_params):
            c = ws_sum.cell(row=r, column=5 + ci,
                            value=sr["analysis"].get(param, "—"))
            c.font = N; c.alignment = C
        ws_sum.row_dimensions[r].height = 20

    col_ws = [22, 22, 28] + [max(14, len(p) + 2) for p in all_params]
    for ci, w in enumerate(col_ws):
        ws_sum.column_dimensions[get_column_letter(2 + ci)].width = w

    for sr in samples_results:
        sname = sr["name"][:28]
        ws    = wb.create_sheet(title=sname)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "2E75B6"
        ws.column_dimensions["A"].width = 3

        ws.merge_cells("B2:G2")
        ws["B2"] = f"Raman Analysis — {sr['name']}"
        ws["B2"].font = Font(name="Calibri", bold=True, color="1F4E79", size=14)
        ws["B2"].alignment = C
        ws.row_dimensions[2].height = 28
        ws.merge_cells("B3:G3")
        ws["B3"] = f"Material: {sr['material']}  |  Group: {sr['group']}  |  Laser: {laser_nm:.0f} nm  |  v2.5"
        ws["B3"].font = Font(name="Calibri", size=10, color="7F7F7F", italic=True)
        ws["B3"].alignment = C

        ws.merge_cells("B5:G5")
        ws.cell(row=5, column=2, value="Analysis Parameters").font = Font(
            name="Calibri", bold=True, color="1F4E79", size=13)
        ws.row_dimensions[5].height = 22
        for ci, h in enumerate(["Parameter", "Value"]):
            cell = ws.cell(row=6, column=2 + ci, value=h)
            cell.font = H; cell.fill = HF; cell.alignment = C; cell.border = brd()
        for ri, (param, val) in enumerate(sr["analysis"].items()):
            r   = 7 + ri
            alt = PatternFill("solid", fgColor="D6E4F0" if ri % 2 == 0 else "EBF3FB")
            for ci in range(2):
                ws.cell(row=r, column=2+ci).fill = alt
                ws.cell(row=r, column=2+ci).border = brd()
            ws.cell(row=r, column=2, value=param).font = B
            ws.cell(row=r, column=2).alignment = L
            ws.cell(row=r, column=3, value=str(val)).font = N
            ws.cell(row=r, column=3).alignment = C
            ws.row_dimensions[r].height = 20

        pk_start = 7 + len(sr["analysis"]) + 2
        ws.merge_cells(f"B{pk_start}:G{pk_start}")
        ws.cell(row=pk_start, column=2, value="Fitted Peaks").font = Font(
            name="Calibri", bold=True, color="1F4E79", size=13)
        ws.row_dimensions[pk_start].height = 22
        pk_headers = ["Peak", "Center (cm-1)", "±σ Center", "FWHM (cm-1)", "±σ FWHM",
                      "Height (a.u.)", "Area (a.u.)", "R²"]
        for ci, h in enumerate(pk_headers):
            cell = ws.cell(row=pk_start+1, column=2+ci, value=h)
            cell.font = H; cell.fill = SF; cell.alignment = C; cell.border = brd()
        for ri, (pname, p) in enumerate(sr["peaks"].items()):
            r   = pk_start + 2 + ri
            alt = PatternFill("solid", fgColor="D6E4F0" if ri % 2 == 0 else "EBF3FB")
            for ci in range(len(pk_headers)):
                ws.cell(row=r, column=2+ci).fill = alt
                ws.cell(row=r, column=2+ci).border = brd()
            ws.cell(row=r, column=2, value=p["name"]).font = B
            ws.cell(row=r, column=2).alignment = L
            if p["found"]:
                c_std  = p.get("center_stderr")
                f_std  = p.get("fwhm_stderr")
                vals   = [
                    round(float(p["center"]), 2),
                    round(float(c_std), 3) if c_std is not None else "—",
                    round(float(p["fwhm"]), 2),
                    round(float(f_std), 3) if f_std is not None else "—",
                    round(float(p["amplitude"]), 1),
                    round(float(p["area"]), 1),
                    round(float(p["r2"]), 4),
                ]
                for ci, v in enumerate(vals):
                    ws.cell(row=r, column=3+ci, value=v).font = N
                    ws.cell(row=r, column=3+ci).alignment = C
            else:
                ws.cell(row=r, column=3, value="Not detected").font = N
            ws.row_dimensions[r].height = 20

        ds = pk_start + 2 + len(sr["peaks"]) + 2
        ws.merge_cells(f"B{ds}:G{ds}")
        ws.cell(row=ds, column=2, value="Spectrum Data").font = Font(
            name="Calibri", bold=True, color="1F4E79", size=13)
        for ci, h in enumerate(["Wavenumber (cm-1)", "Raw Intensity",
                                  "Baseline", "Corrected", "Normalised (0-1)"]):
            cell = ws.cell(row=ds+1, column=2+ci, value=h)
            cell.font = H; cell.fill = HF; cell.alignment = C
        corr = sr["corrected"]
        norm = corr / corr.max() if corr.max() > 0 else corr
        for i, (w, r_, b, co, n) in enumerate(
                zip(sr["wn"], sr["intensity"], sr["baseline"], corr, norm)):
            ri2 = ds + 2 + i
            for ci, v in enumerate([round(float(w),3), round(float(r_),3),
                                      round(float(b),3), round(float(co),3),
                                      round(float(n),5)]):
                ws.cell(row=ri2, column=2+ci, value=v)
        for c, w_ in [(2,22),(3,18),(4,18),(5,22),(6,18),(7,16)]:
            ws.column_dimensions[get_column_letter(c)].width = w_

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════
#  STREAMLIT APP
# ══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Raman Analyzer — 2D Materials",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

with st.sidebar:
    st.title("🔬 Raman Analyzer")
    st.caption("Graphene · TMDs · h-BN · BP · MXene · v2.5")
    st.divider()

    st.subheader("🔴 Laser Wavelength")
    laser_nm = st.number_input("Wavelength (nm)", min_value=400.0,
                                max_value=1100.0, value=532.0, step=1.0)
    c1, c2, c3, c4 = st.columns(4)
    presets = {488: c1, 532: c2, 633: c3, 785: c4}
    for nm, col in presets.items():
        if col.button(str(nm), use_container_width=True):
            st.session_state["laser_nm"] = float(nm)
            st.rerun()
    if "laser_nm" in st.session_state:
        laser_nm = st.session_state["laser_nm"]

    st.subheader("📉 Baseline Correction")
    baseline_method = st.selectbox("Method", ["als", "linear"])
    als_lam = st.number_input("ALS lambda", value=1e5, min_value=1e2,
                               max_value=1e9, format="%.0e",
                               disabled=(baseline_method != "als"))
    als_p   = st.number_input("ALS p", value=0.001, min_value=1e-4,
                               max_value=0.5, format="%.4f",
                               disabled=(baseline_method != "als"))
    st.divider()

    st.subheader("📥 Step 1 — Download Template")
    n_samples = st.number_input("Number of samples", min_value=1,
                                 max_value=50, value=3, step=1)
    st.download_button(
        label="⬇️  Download Excel Template",
        data=make_template(int(n_samples)),
        file_name=f"raman_template_{int(n_samples)}samples.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("📌 Open the READ ME FIRST sheet before filling data.")

    st.subheader("📤 Step 2 — Upload Filled Template")
    uploaded = st.file_uploader(
        "Upload Excel file", type=["xlsx"],
        help="Each sheet = one sample. Fill sample name in the yellow cell B2."
    )

st.title("Raman Spectrum Analyzer — 2D Materials")
st.caption("Graphene / rGO / GO / TMDs / h-BN / Black Phosphorus / MXene  •  v2.5")

if uploaded is None:
    st.info("👈  Download the template (Step 1), fill it in, then upload it (Step 2).")
    with st.expander("ℹ️  Quick Guide"):
        st.markdown("""
**Step 1** — Set number of samples → click *Download Excel Template*

**Step 2** — Open the downloaded file, read the **READ ME FIRST** sheet

**Step 3** — In each sample sheet:
- Cell **B2** (yellow): type your sample name *(no / \\ : ? * [ ] — max 28 chars)*
- From **row 4** onward: Wavenumber in column B, Intensity in column C

**Step 4** — Upload the filled file → configure material for each sample → **RUN ANALYSIS**

---
**v2.4 / v2.5 features now in UI:**
- 🔬 D* band quantification (I_D*/I_G, C/O ratio proxy) [Lee 2021]
- ⚠️ B-doping fingerprint auto-detection [Kim 2012]
- 📐 Fitting uncertainty — center ± σ, FWHM ± σ [pcov]
- ⚡ Doping level estimator (n/p-type + carrier density) [Pisana 2007]
- 🔀 Refined stage boundary (FWHM(G) + A_D/A_G) [Wu 2018]
- 📊 Batch statistics + ratio heatmap [Roadmap #10]
        """)
    st.stop()

try:
    xl     = pd.ExcelFile(uploaded)
    sheets = [s for s in xl.sheet_names if not str(s).upper().startswith("READ")]
except Exception as e:
    st.error(f"Cannot read Excel file: {e}")
    st.stop()

if not sheets:
    st.error("No data sheets found. Make sure your file has sheets with sample data.")
    st.stop()


def read_sample_name_from_sheet(xl, sheet):
    """Read custom name from cell B2 (index [1,1]); fall back to sheet title."""
    try:
        df_peek = xl.parse(sheet, header=None, nrows=3)
        if df_peek.shape[0] > 1 and df_peek.shape[1] > 1:
            val = str(df_peek.iloc[1, 1]).strip()
            if val and val.lower() not in ("nan", "", "none"):
                return val
    except Exception:
        pass
    return sheet


st.subheader(f"📋 Step 3 — Configure {len(sheets)} sample(s)")
st.caption("Material group and specific material are selected **outside the form** — "
           "so the material list updates instantly when you change the group.")

group_options = list(MATERIAL_GROUPS.keys())

if "sample_configs" not in st.session_state:
    st.session_state.sample_configs = {}

name_errors   = []
final_configs = []

for i, sheet in enumerate(sheets):
    default_name = read_sample_name_from_sheet(xl, sheet)
    key_prefix   = f"s{i}"

    with st.container():
        st.markdown(f"**Sample {i+1} — Sheet: `{sheet}`**")
        col_name, col_group, col_mat = st.columns([2, 3, 3])

        sample_name = col_name.text_input(
            "Sample name", value=default_name, key=f"{key_prefix}_name"
        )
        valid, err_msg = validate_sample_name(sample_name)
        if not valid:
            col_name.error(f"⚠️ {err_msg}")
            name_errors.append(err_msg)

        group = col_group.selectbox(
            "Material group", group_options, key=f"{key_prefix}_group"
        )
        materials = MATERIAL_GROUPS[group]
        material  = col_mat.selectbox(
            "Material", materials, key=f"{key_prefix}_mat"
        )

        final_configs.append({
            "sheet": sheet, "name": sample_name.strip(),
            "group": group, "material": material,
        })

    if i < len(sheets) - 1:
        st.divider()

st.markdown("")
run_disabled = len(name_errors) > 0
if run_disabled:
    st.warning("Fix the sample name errors above before running analysis.")

if st.button("▶  RUN ANALYSIS", type="primary",
              use_container_width=True, disabled=run_disabled):
    st.session_state["run"]     = True
    st.session_state["configs"] = final_configs

if not st.session_state.get("run"):
    st.stop()

# ── Run ───────────────────────────────────────────────────
configs = st.session_state["configs"]
st.subheader("🔄 Processing...")
progress = st.progress(0)
samples_results, errors = [], []

for i, cfg in enumerate(configs):
    try:
        wn, intensity = _parse_sheet_data(xl, cfg["sheet"])
        sort_idx      = np.argsort(wn)
        wn, intensity = wn[sort_idx], intensity[sort_idx]

        mat_key = _material_key(cfg["group"], cfg["material"])
        peaks, analysis, corrected, baseline_arr, validation_report = run_analysis(
            wn, intensity, mat_key, cfg["group"],
            laser_nm, baseline_method, als_lam, als_p
        )
        samples_results.append({
    **cfg, "mat_key": mat_key,
    "wn": wn, "intensity": intensity,
    "baseline": baseline_arr, "corrected": corrected,
    "peaks": peaks, "analysis": analysis,
    "validation": validation_report,
})
    except Exception as e:
        errors.append(f"{cfg['sheet']}: {e}")
    progress.progress((i + 1) / len(configs))

for err in errors:
    st.warning(f"⚠️  {err}")

if not samples_results:
    st.error("No samples could be processed.")
    st.session_state["run"] = False
    st.stop()

st.success(f"✅  {len(samples_results)} sample(s) analysed!")

# ── Result tabs ───────────────────────────────────────────
tab_labels = (
    ["📈 All Spectra", "🔍 Peak Fits", "📋 Results Table", "📊 Batch Statistics"]
    + [f"📄 {sr['name']}" for sr in samples_results]
)
tab_objects = st.tabs(tab_labels)
palette = ["#4fc3f7","#69db7c","#ff6b6b","#ffa94d",
           "#cc99ff","#ffd43b","#a9e34b","#f783ac"]

# ── Tab 0: All Spectra ────────────────────────────────────
with tab_objects[0]:
    st.subheader("All Spectra — Baseline Corrected")
    fig_all = go.Figure()
    for idx, sr in enumerate(samples_results):
        fig_all.add_trace(go.Scatter(
            x=sr["wn"], y=sr["corrected"],
            name=sr["name"],
            line=dict(color=palette[idx % len(palette)], width=1.5),
        ))
    fig_all.update_layout(
        template="plotly_dark", height=450,
        xaxis_title="Raman Shift (cm⁻¹)",
        yaxis_title="Intensity (a.u.)",
        legend=dict(orientation="h", y=1.08),
    )
    st.plotly_chart(fig_all, use_container_width=True)

# ── Tab 1: Peak Fits ──────────────────────────────────────
with tab_objects[1]:
    st.subheader("Peak Fits")
    sel_name = st.selectbox("Select sample", [sr["name"] for sr in samples_results],
                             key="peak_sel")
    sr_sel   = next(s for s in samples_results if s["name"] == sel_name)
    colors   = _make_peak_colors(sr_sel["mat_key"])
    found_pk = [(k, p) for k, p in sr_sel["peaks"].items()
                if p["found"] and len(p["x"]) > 0]
    if not found_pk:
        st.warning("No peaks detected.")
    else:
        cols_pk = st.columns(min(len(found_pk), 3))
        for idx, (key, p) in enumerate(found_pk):
            color = colors.get(key, "#cccccc")
            mask  = (sr_sel["wn"] >= p["x"][0]) & (sr_sel["wn"] <= p["x"][-1])
            xd, yd = sr_sel["wn"][mask], sr_sel["corrected"][mask]

            # v2.4 Feature #3: show stderr if available
            ctr_str  = f"{p['center']:.1f}"
            if p.get("center_stderr") is not None:
                ctr_str += f" ± {p['center_stderr']:.1f}"
            fwhm_str = f"{p['fwhm']:.1f}"
            if p.get("fwhm_stderr") is not None:
                fwhm_str += f" ± {p['fwhm_stderr']:.1f}"

            fig_pk = go.Figure()
            fig_pk.add_trace(go.Scatter(
                x=xd, y=yd, mode="markers",
                marker=dict(size=4, color="#cdd6f4", opacity=0.6), name="Data"
            ))
            fig_pk.add_trace(go.Scatter(
                x=p["x"], y=p["y_fit"],
                line=dict(color=color, width=2.5),
                fill="tozeroy",
                fillcolor=f"rgba({','.join(str(int(color.lstrip('#')[j:j+2],16)) for j in (0,2,4))},0.15)",
                name="Fit"
            ))
            split   = " [dual-L]"  if p.get("is_split_2D")    else ""
            deconv  = " [deconv]"  if p.get("is_deconvolved") else ""
            fig_pk.update_layout(
                title=dict(
                    text=(
                        f"<b style='color:{color}'>{p['name']}{split}{deconv}</b><br>"
                        f"Center: {ctr_str} cm⁻¹  |  FWHM: {fwhm_str} cm⁻¹  |  R²={p['r2']:.3f}"
                    ),
                    font=dict(size=10)
                ),
                template="plotly_dark", height=290, showlegend=False,
                margin=dict(t=85, b=40, l=40, r=10),
                xaxis_title="Raman Shift (cm⁻¹)",
                yaxis_title="Intensity",
            )
            cols_pk[idx % 3].plotly_chart(fig_pk, use_container_width=True)

# ── Tab 2: Results Table ──────────────────────────────────
with tab_objects[2]:
    st.subheader("Results Summary")
    all_pk, seen_k = [], set()
    for sr in samples_results:
        for k in sr["analysis"]:
            if k not in seen_k:
                all_pk.append(k); seen_k.add(k)
    rows = []
    for sr in samples_results:
        row = {"Sample": sr["name"], "Material": sr["material"]}
        for k in all_pk:
            row[k] = sr["analysis"].get(k, "—")
        rows.append(row)
    df_results = pd.DataFrame(rows)
    st.dataframe(df_results, hide_index=True, use_container_width=True)

# ── Tab 3: Batch Statistics + Heatmap (Roadmap #10) ──────
with tab_objects[3]:
    st.subheader("📊 Batch Statistics — Ratio Heatmap")

    graphene_samples = [sr for sr in samples_results if sr["mat_key"] in ("graphene", "mxene")]

    if len(graphene_samples) < 2:
        st.info("Batch statistics require ≥ 2 graphene/sp² carbon samples. "
                "Non-graphene materials are excluded from ratio heatmap.")
    else:
        df_num = _build_numeric_df(graphene_samples)

        if df_num.empty or df_num.shape[1] == 0:
            st.warning("No numeric ratio data available for the loaded samples.")
        else:
            # ── Heatmap ───────────────────────────────────
            st.markdown("#### Ratio Heatmap — all samples × all numeric ratios")
            st.caption(
                "Colour is **column-normalised** (low→red, high→green). "
                "Cell values show raw ratio numbers."
            )
            fig_hm = _ratio_heatmap(df_num)
            st.plotly_chart(fig_hm, use_container_width=True)

            # ── Per-ratio bar chart ───────────────────────
            st.markdown("#### Per-Ratio Bar Chart")
            ratio_choices = [c for c in df_num.columns if c in HEATMAP_RATIO_KEYS]
            if not ratio_choices:
                ratio_choices = df_num.columns.tolist()

            sel_ratio = st.selectbox(
                "Select ratio to plot", ratio_choices, key="batch_ratio_sel"
            )
            fig_bar = _ratio_bar_chart(df_num, sel_ratio)
            st.plotly_chart(fig_bar, use_container_width=True)

            # ── Descriptive stats table ───────────────────
            st.markdown("#### Descriptive Statistics")
            stats = df_num.describe().T.round(4)
            stats.index.name = "Ratio"
            st.dataframe(stats, use_container_width=True)

            # ── CSV download ──────────────────────────────
            csv_bytes = _numeric_df_to_csv(df_num)
            st.download_button(
                label="⬇️  Download Numeric Ratios (CSV)",
                data=csv_bytes,
                file_name=f"raman_batch_ratios_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # All materials: show text stats for non-graphene too
    if len(samples_results) > len(graphene_samples):
        with st.expander("🔍 Non-graphene sample parameters"):
            other = [sr for sr in samples_results if sr["mat_key"] not in ("graphene","mxene")]
            for sr in other:
                st.markdown(f"**{sr['name']}** ({sr['material']})")
                st.json(sr["analysis"])

# ── Per-sample tabs ───────────────────────────────────────
for ti, sr in enumerate(samples_results):
    with tab_objects[4 + ti]:
        st.subheader(f"{sr['name']} — {sr['material']}")

        fig_s = make_subplots(rows=2, cols=1,
                               subplot_titles=("Raw + Baseline", "Corrected + peak positions"),
                               vertical_spacing=0.12)
        fig_s.add_trace(go.Scatter(x=sr["wn"], y=sr["intensity"],
                                    name="Raw", line=dict(color="#4fc3f7", width=1.2)),
                         row=1, col=1)
        fig_s.add_trace(go.Scatter(x=sr["wn"], y=sr["baseline"],
                                    name="Baseline",
                                    line=dict(color="#ff6b6b", width=1.5, dash="dash")),
                         row=1, col=1)
        fig_s.add_trace(go.Scatter(x=sr["wn"], y=sr["corrected"],
                                    name="Corrected",
                                    line=dict(color="#69db7c", width=1.4),
                                    fill="tozeroy",
                                    fillcolor="rgba(105,219,124,0.07)"),
                         row=2, col=1)
        colors_s = _make_peak_colors(sr["mat_key"])
        for key, p in sr["peaks"].items():
            if p["found"]:
                col = colors_s.get(key, "gray")
                fig_s.add_vline(x=p["center"], line_width=0.8,
                                 line_dash="dot", line_color=col, row=2, col=1)
                fig_s.add_annotation(
                    x=p["center"], y=sr["corrected"].max() * 0.88,
                    text=f"<b>{p['name']}</b>",
                    font=dict(color=col, size=10),
                    showarrow=False, row=2, col=1
                )
        fig_s.update_layout(
            height=580, template="plotly_dark",
            title=dict(text=f"{sr['name']} | laser={laser_nm:.0f} nm  [v2.5]",
                        font=dict(size=13)),
            legend=dict(orientation="h", y=1.06),
        )
        fig_s.update_xaxes(title_text="Raman Shift (cm⁻¹)", row=2, col=1)
        fig_s.update_yaxes(title_text="Intensity (a.u.)")
        st.plotly_chart(fig_s, use_container_width=True)

        # ── Analysis parameters table ─────────────────
        st.subheader("Analysis Parameters  (v2.5)")

        # Highlight special fields
        def _row_style(row):
            if "B-Doping" in str(row.get("Parameter", "")) and "YES" in str(row.get("Value", "")):
                return ["background-color: #3d1a1a; color: #ff9999"] * 2
            if "Stage Refined" in str(row.get("Parameter", "")) and "Stage 2" in str(row.get("Value", "")):
                return ["background-color: #2a2a1a; color: #ffd43b"] * 2
            if "Doping Type" in str(row.get("Parameter", "")):
                return ["background-color: #1a2a2a; color: #69db7c"] * 2
            return [""] * 2

        df_params = pd.DataFrame(
            [{"Parameter": k, "Value": v} for k, v in sr["analysis"].items()]
        )
        st.dataframe(
            df_params.style.apply(_row_style, axis=1),
            hide_index=True, use_container_width=True
        )

        # ── Fitted peak table ─────────────────────────
        st.subheader("Fitted Peaks (with ± uncertainty)")
        pk_rows = []
        for k, p in sr["peaks"].items():
            if p["found"]:
                c_std = p.get("center_stderr")
                f_std = p.get("fwhm_stderr")
                pk_rows.append({
                    "Peak":         p["name"],
                    "Center (cm⁻¹)": f"{p['center']:.2f}" + (f" ± {c_std:.2f}" if c_std else ""),
                    "FWHM (cm⁻¹)":   f"{p['fwhm']:.2f}"  + (f" ± {f_std:.2f}" if f_std else ""),
                    "Height":        f"{p['amplitude']:.1f}",
                    "Area":          f"{p['area']:.1f}",
                    "R²":            f"{p['r2']:.4f}",
                    "Notes":         ("dual-L" if p.get("is_split_2D") else "") +
                                     ("deconv" if p.get("is_deconvolved") else ""),
                })
        if pk_rows:
            st.dataframe(pd.DataFrame(pk_rows), hide_index=True, use_container_width=True)
        else:
            st.info("No peaks detected in this sample.")

# ── Excel export ──────────────────────────────────────────
st.divider()
st.subheader("📊 Download Full Excel Report  (v2.5)")
excel_bytes = build_excel_report(samples_results, laser_nm)
st.download_button(
    label="⬇️  Download Excel Report",
    data=excel_bytes,
    file_name=f"raman_report_v25_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    type="primary",
)
